from openpyxl.styles import Alignment, PatternFill, Font, Color
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl
import calendar

# Set the first day as Sunday

calendar.setfirstweekday(firstweekday=6)
c = calendar.TextCalendar(calendar.SUNDAY)
# Create a job

wb = openpyxl.Workbook()

# Traveling for 12 months


for i in range(1, 13):
# Add worksheets
	sheet = wb.create_sheet(index=0, title=str(i) + 'month')
	sht = wb.get_active_sheet()
	sht.show_grid_lines = False
	step = 0
	
	side = openpyxl.styles.Side(style=None, color='ffffff')
	no_border = openpyxl.styles.borders.Border(
    left=side, 
    right=side, 
    top=side, 
    bottom=side,
	)
	
	# Get the exact date and time
	for j in range(len(calendar.monthcalendar(2019, i))):
		
		sheet.row_dimensions[j + 3 + step + 1].height = 45

		for k in range(len(calendar.monthcalendar(2019, i)[j])):
			sheet.column_dimensions[get_column_letter(k+1)].width = 20
			value = calendar.monthcalendar(2019, i)[j][k]
			
			# Change 0 to null
			cell = sheet.cell(row=j + 3 + step, column=k + 1)
			
			if value == 0:
				value = ''
				cell.value
			else:
				cell.value = value
				# Setting fonts
				cell.font = Font(u'Microsoft YaHei', size=11)
				align = Alignment(horizontal='left', vertical='center')
				fill = PatternFill("solid", fgColor="f2f2f2")

				cell.fill = fill
				cell.border = no_border
				# Cell Text Settings, Right Alignment, Vertical Centralization
				cell.alignment = align
				# Cell Fill Color Property Settings

		step += 1		

	# Add Weekly Information Line
	days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
	align = Alignment(horizontal='center', vertical='center')
	bottom_border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.Side(style='thin'))
	num = 0

	for k3 in range(1, 8):
		sheet.cell(row=2, column=k3).value = days[k3-1]
		sheet.cell(row=2, column=k3).alignment = align
		sheet.cell(row=2, column=k3).font = Font(u'Rockwell', size=11)
		sheet.cell(row=2, column=k3).border = bottom_border
		sheet.cell(row=1, column=k3).fill = PatternFill("solid", fgColor='ff8533')

	#remove all borders
	'''
	for x in range(100):
		for y in range(100):
			sheet.cell(row=x+1, column=y+1).border=no_border
   '''

	# Added year and month
	fill = PatternFill("solid", fgColor="ff8000")
	sheet.cell(row=1, column=1).value = str(i) + 'month 2020'
	# Setting year and month text properties
	sheet.cell(row=1, column=1).font = Font(u'Rockwell', size=35)
	#sheet.cell(row=4, column=1).font = Font(u'Microsoft YaHei', size=16, bold=True, color='FF7887')
	#sheet.cell(row=3, column=1).alignment = align
	#sheet.cell(row=4, column=1).alignment = align


# Save Documents

wb.save('Aidou Calendar.xlsx')