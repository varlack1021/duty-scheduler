import xlrd
import calendar
import xlsxwriter
from datetime import datetime
from openpyxl.styles import Font
from openpyxl import load_workbook
from pathlib import Path

#list of cell fill in colors for different months
month_styles = {1: '0099ff', 2: '66c2ff', 3: '8cd98c', 4: '339933', 5: '00802b', 6: 'ffdd99',
				7: 'ffbb33', 8: 'ff9933', 9: 'e65c00', 10: 'cc0000', 11: '990000', 12: '0066cc'}

EXCEL_FOLDER_PATH = 'backend/excel_files/'
# Set the first day as Sunday

class Excel_Calendar:
	
	def __init__(self, filename):
		self.filename = filename + '.xlsx'
		self.path = Path(EXCEL_FOLDER_PATH) / self.filename

	def create_excel_calendar(self, date_range):
		calendar.setfirstweekday(firstweekday=6)
		workbook = xlsxwriter.Workbook(self.path)

		for date in sorted(list(date_range)):
			# Add worksheets
			year = date.year
			month = date.month
			month_name = calendar.month_name[date.month]
			worksheet = workbook.add_worksheet(month_name)
			worksheet.set_column('A:H',20)
			worksheet.set_row(1, 20)
			worksheet.set_row(0, 55)
			worksheet.hide_gridlines(2)
			step = 0

			# Add cell styles
			date_format = workbook.add_format()
			day_format = workbook.add_format()
			title_format = workbook.add_format()
			gray_fill = workbook.add_format()
			total_duty_date_format = workbook.add_format()

			gray_fill.set_bg_color('e6e6e6')
			gray_fill.set_align('left')
			
			#day number
			date_format.set_align('left')
			date_format.set_align('top')
			
			#day of the week text i.e Sunday, Monday
			day_format.set_font_name('Rockwell')
			day_format.set_font_size(11)
			day_format.set_bottom()
			day_format.set_align('vcenter')
			day_format.set_align('center')
			
			title_format.set_font_size(35)
			title_format.set_bg_color(month_styles[month])
			title_format.set_font_name('Rockwell')
			title_format.set_align('vcenter')

			total_duty_date_format.set_align('left')
			total_duty_date_format.set_align('top')
			total_duty_date_format.set_align('vcenter')
			total_duty_date_format.set_align('center')

			# Get the exact date and time
			fill_in = False

			for j in range(len(calendar.monthcalendar(year, month))):
				

				worksheet.set_row(j + 3 + step, 45)

				for k in range(len(calendar.monthcalendar(year, month)[j])):

					value = calendar.monthcalendar(year, month)[j][k]
					
					if value != 0:
						worksheet.write(j + 2 + step, k, value, date_format)

						if fill_in:
							worksheet.write(j + 2 + step, k, value, gray_fill)
							worksheet.write(j + 3 + step, k, '', gray_fill)

				fill_in = True if not fill_in else False

				step += 1		

			# Add Weekly Information Line
			
			days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
			
			for k3 in range(7):

				worksheet.write(1, k3, days[k3], day_format)
				worksheet.write(0, k3, '', title_format)
			
			month_name = calendar.month_name[month]
			worksheet.write(0, 0, "{}  {}".format(month_name, year), title_format)
		
		# Add Total Duty Dates Headers
		worksheet = workbook.add_worksheet("Total Duty Days")
		worksheet.set_column('A:C',20)

		headers = ['Name', 'Total Weekdays', 'Total Weekends']
		for i in range(len(headers)):
			worksheet.write(0, i, headers[i], total_duty_date_format)
		# Save Documents
		workbook.close()

	def write_to_excelsheet(self, duty_dates, total_days, date_range):
	    '''
	    Creates and writes to excel sheet
	    The algorithm iterates through the cells in the excel sheet that have a date.
	    The algo checks if the date is inlcuded in scheduled duty dates
	    '''
	    self.create_excel_calendar(date_range)
	    path = self.path
	    #wb = xlrd.open_workbook(loc)

	    write_wb = load_workbook(self.path)
	    read_wb = xlrd.open_workbook(self.path)

	    for month_number in [date.month for date in date_range]:
	        datetime_object = datetime.strptime(str(month_number), "%m")
	        month_name = datetime_object.strftime('%B')
	        write_sheet = write_wb[month_name]
	        read_sheet = read_wb.sheet_by_name(month_name)

	        for row in range(2, write_sheet.max_row, 2):
	            for col in range(write_sheet.max_column):
	                day = read_sheet.cell_value(row, col)
	                if day in duty_dates[month_number].keys() and day:
	                    row_num = row + 1
	                    name = duty_dates[month_number][day]
	                    cell = write_sheet.cell(row_num + 1, col + 1)
	                    cell.value = name

	                    if cell.value == 'Unable to Assign':
	                    	cell.font = Font(color='eb0c0c')

	    # Add total duty days sat per person sheet
	    write_sheet = write_wb['Total Duty Days']

	    for index, name in enumerate(total_days):
	    	write_sheet.cell(index + 2, 1).value = name
	    	write_sheet.cell(index + 2, 2).value = total_days[name]['weekdays']
	    	write_sheet.cell(index + 2, 3).value = total_days[name]['weekends']

	    write_wb.save(self.path)

	def get_duty_dates_from_sheet(self, name):
	    duty_dates = {}
	    wb = xlrd.open_workbook(self.path)
	    
	    for sheet in wb.sheets():

	        if any(sheet.row_values(5)):
	            month_number = list(calendar.month_name).index(sheet.name)
	            duty_dates[month_number] = []
	            
	            for row in range(2, sheet.nrows):

	                row_values = sheet.row_values(row)
	                if name in row_values:
	                    columns = [i for i, x in enumerate(row_values) if x == name]
	                    days = [sheet.cell(row-1, x).value for x in columns]
	                    duty_dates[month_number].extend(days)
	    
	    if duty_dates[list(duty_dates.keys())[0]][0] > 31:        
	        duty_dates = convert_excel_date(duty_dates)

	    return duty_dates